import requests
from fake_useragent import UserAgent
from lxml import etree
import time
import re
from openpyxl import Workbook, load_workbook

ua = UserAgent(verify_ssl=False)
domain_url = "https://book.douban.com/tag/"

def getBookTagList():
    headers = {
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
        "Accept-Encoding": "gzip, deflate, br",
        "Accept-Language": "zh,en-US;q=0.9,en;q=0.8,zh-TW;q=0.7,zh-CN;q=0.6",
        "Connection": "keep-alive",
        # "Cookie": "bid=SXZeHMYH16M; douban-fav-remind=1; gr_user_id=7d4342b5-f8cc-4e60-b9ca-3a496d3ae3b8; gr_session_id_22c937bbd8ebd703f2d8e9445f7dfd03=04946e2a-49c4-4cd4-8c44-b7506ca0b6ec; gr_cs1_04946e2a-49c4-4cd4-8c44-b7506ca0b6ec=user_id%3A0; __utma=30149280.363311878.1550211797.1550211797.1550929438.2; __utmc=30149280; __utmz=30149280.1550929438.2.2.utmcsr=baidu|utmccn=(organic)|utmcmd=organic; __utma=81379588.358812136.1550929438.1550929438.1550929438.1; __utmc=81379588; __utmz=81379588.1550929438.1.1.utmcsr=baidu|utmccn=(organic)|utmcmd=organic; _pk_ref.100001.3ac3=%5B%22%22%2C%22%22%2C1550929438%2C%22https%3A%2F%2Fwww.baidu.com%2Flink%3Furl%3D4-jiW7eEoxuKVWnPf886mJDeo4-kIzmMu5D3fN-9TMmJJglliEIQ_9boBZ3cYYX1%26wd%3D%26eqid%3D8f274b1a00000bae000000055c714e19%22%5D; _pk_ses.100001.3ac3=*; ap_v=0,6.0; gr_session_id_22c937bbd8ebd703f2d8e9445f7dfd03_04946e2a-49c4-4cd4-8c44-b7506ca0b6ec=true; __yadk_uid=KqHoYC6gRyVuMmKHQCmXNZF07f05vxdn; _vwo_uuid_v2=D86724BDF751C0E31A774A01ADE1EEDD9|e5bddedca8f2f86974aad133a5106cbb; __utmt_douban=1; __utmt=1; _pk_id.100001.3ac3=502f0d3c518430a7.1550929438.1.1550931001.1550929438.; __utmb=30149280.19.10.1550929438; __utmb=81379588.19.10.1550929438",
        "Host": "book.douban.com",
        "Upgrade-Insecure-Requests": "1",
        "User-Agent": ua.random# "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.109 Safari/537.36",
    }
    r = requests.get(url=domain_url,headers=headers)
    html = etree.HTML(r.text)
    tags = html.xpath("//table[@class='tagCol']/tbody/tr/td/a/text()")
    return tags

def getBooks(url):
    headers = {
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
        "Accept-Encoding": "gzip, deflate, br",
        "Accept-Language": "zh,en-US;q=0.9,en;q=0.8,zh-TW;q=0.7,zh-CN;q=0.6",
        "Connection": "keep-alive",
        # "Cookie": "bid=SXZeHMYH16M; douban-fav-remind=1; gr_user_id=7d4342b5-f8cc-4e60-b9ca-3a496d3ae3b8; __utma=30149280.363311878.1550211797.1550211797.1550929438.2; __utmc=30149280; __utmz=30149280.1550929438.2.2.utmcsr=baidu|utmccn=(organic)|utmcmd=organic; __utma=81379588.358812136.1550929438.1550929438.1550929438.1; __utmc=81379588; __utmz=81379588.1550929438.1.1.utmcsr=baidu|utmccn=(organic)|utmcmd=organic; _pk_ref.100001.3ac3=%5B%22%22%2C%22%22%2C1550929438%2C%22https%3A%2F%2Fwww.baidu.com%2Flink%3Furl%3D4-jiW7eEoxuKVWnPf886mJDeo4-kIzmMu5D3fN-9TMmJJglliEIQ_9boBZ3cYYX1%26wd%3D%26eqid%3D8f274b1a00000bae000000055c714e19%22%5D; _pk_ses.100001.3ac3=*; ap_v=0,6.0; __yadk_uid=KqHoYC6gRyVuMmKHQCmXNZF07f05vxdn; _vwo_uuid_v2=D86724BDF751C0E31A774A01ADE1EEDD9|e5bddedca8f2f86974aad133a5106cbb; __utmt_douban=1; __utmt=1; _pk_id.100001.3ac3=502f0d3c518430a7.1550929438.1.1550932734.1550929438.; __utmb=30149280.32.10.1550929438; __utmb=81379588.32.10.1550929438",
        "Host": "book.douban.com",
        "Referer": url.replace("S","").encode(),
        "Upgrade-Insecure-Requests": "1",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.109 Safari/537.36",
    }
    r = requests.get(url=url,headers=headers)
    html = etree.HTML(r.text)
    bookNames = html.xpath("//div[@class='info']/h2/a")
    if len(bookNames) == 0:
        return
    bookDetails = html.xpath("//div[@class='info']/div[@class='pub']")
    bookRatingNums = html.xpath("//div[@class='info']//span[@class='rating_nums']")
    bookEvaluateNums = html.xpath("//div[@class='info']//span[@class='pl']")
    books = []
    for index,bookName in enumerate(bookNames):
        book = {}
        book["name"] = bookName.text.strip()
        book["detail"] = bookDetails[index].text.strip()
        book["rating"] = bookRatingNums[index].text
        book["num"] = re.sub("\D","",bookEvaluateNums[index].text.strip())
        books.append(book)
    print(books)
    time.sleep(1)
    return books
    
def handler_url(tag):
    print("当前工作标签："+tag)
    wb = load_workbook("doubandushu.xlsx")
    ws = wb.create_sheet("Mysheet")           #创建一个sheet
    ws.title = tag
    ws['A1'] = "书名"
    ws['B1'] = "作者/出版社/出版时间/定价"
    ws['C1'] = "评分"
    ws['D1'] = "评论人数"
    num = 0
    while(True):
        url = domain_url+tag+"?start="+str(num)+"&type=S"
        books = getBooks(url)
        if books == None:
            break
        else:
            for book in books:
                if int(book["num"]) > 1000:
                    ws.append([book["name"],book["detail"],book["rating"],book["num"]])
            num = num+20
    wb.save("doubandushu.xlsx")

bookTagList = getBookTagList()
wb = Workbook()
wb.save("doubandushu.xlsx")
for bookTag in bookTagList:
   handler_url(bookTag)
