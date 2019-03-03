import requests
from fake_useragent import UserAgent
from lxml import etree
from openpyxl import Workbook, load_workbook
import time

domain_url = "https://coding.imooc.com/"
ua = UserAgent(verify_ssl=False)
wb = Workbook()
workbook_name = "moocList.xlsx"

def getClassification():
    headers = {
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
        "Accept-Encoding": "gzip, deflate, br",
        "Accept-Language": "zh,en-US;q=0.9,en;q=0.8,zh-TW;q=0.7,zh-CN;q=0.6",
        "Cache-Control": "max-age=0",
        "Connection": "keep-alive",
        "Host": "coding.imooc.com",
        "Referer": "https://www.imooc.com/course/list",
        "Upgrade-Insecure-Requests": "1",
        "User-Agent": ua.random
    # "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.119 Safari/537.36",
    }
    response = requests.get(url=domain_url, headers=headers)
    response.encoding = "utf-8"
    html = etree.HTML(response.text)
    lists = html.xpath("//div[@class='shizhan-header-nav']/div/a/text()")[1:]
    lists_url = html.xpath("//div[@class='shizhan-header-nav']/div/a/@href")[1:]
    list_data = [(list, url) for list, url in zip(lists, lists_url)]
    for x in list_data:
        classify_name = x[0]
        classify_url = domain_url+x[1]
        ws = wb.create_sheet(classify_name)
        getSubclassification(classify_url,ws)
        time.sleep(10)

def getSubclassification(url,ws):
    headers = {
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
        "Accept-Encoding": "gzip, deflate, br",
        "Accept-Language": "zh,en-US;q=0.9,en;q=0.8,zh-TW;q=0.7,zh-CN;q=0.6",
        "Connection": "keep-alive",
        "Host": "coding.imooc.com",
        "Referer": "https://coding.imooc.com/",
        "Upgrade-Insecure-Requests": "1",
        "User-Agent": ua.random
    }
    response = requests.get(url=url,headers=headers)
    response.encoding = "utf-8"
    html = etree.HTML(response.text)
    sub_lists = html.xpath("//div[@class='shizhan-skill clearfix']/a/text()")[1:]
    sub_url_lists = html.xpath("//div[@class='shizhan-skill clearfix']/a/@href")[1:]
    sub_datas = [(list, url) for list, url in zip(sub_lists, sub_url_lists)]
    for y in sub_datas:
        sub_name = ["分类："]
        sub_name.append(y[0].strip())
        print(sub_name)
        sub_url = domain_url+y[1]
        ws.append(sub_name)
        ws.append(["课程名称","老师","难度等级","学习人数","评分","介绍","价格"])
        courses_list = getCourses(sub_url)
        for course in courses_list:
            ws.append([course[0],course[1],course[2],course[3],course[4],course[5],course[6]])

def getCourses(url):
    courses_list = []
    page = 1
    headers = {
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
        "Accept-Encoding": "gzip, deflate, br",
        "Accept-Language": "zh,en-US;q=0.9,en;q=0.8,zh-TW;q=0.7,zh-CN;q=0.6",
        "Cache-Control": "max-age=0",
        "Connection": "keep-alive",
        "Host": "coding.imooc.com",
        "Referer": "https://coding.imooc.com/",
        "Upgrade-Insecure-Requests": "1",
        "User-Agent": ua.random # "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.119 Safari/537.36",
    }
    while(True):
        url = url + "&sort=2&unlearn=0&page=" + str(page)
        response = requests.get(url=url,headers=headers)
        response.encoding = "utf-8"
        html = etree.HTML(response.text)
        names = html.xpath("//div[@class='shizhan-intro-box']/p[1]/text()")
        if len(names) == 0:
            break
        teachers = html.xpath("//div[@class='lecturer-info']/span/text()")
        courses_levels = html.xpath("//div[@class='shizhan-info']/span[1]/text()")
        learn_sum = html.xpath("//div[@class='shizhan-info']/span[2]/text()")
        rating_num = html.xpath("//div[@class='shizhan-info']/span[3]/text()")
        intros = html.xpath("//p[@class='shizan-desc']/text()")
        prices = html.xpath("//div[@class='shizhan-info-bottom']/div[1]/*[1]/text()")
        data = [(name.strip(),teacher.strip(),level.strip(),sum.strip(),rating.replace("评价：","").strip(),intro.strip(),price.strip()) for name,teacher,level,sum,rating,intro,price in zip(names, teachers, courses_levels,learn_sum,rating_num,intros,prices )]
        print(data)
        courses_list = courses_list+data
        page = page + 1
    return courses_list

getClassification()
wb.save(workbook_name)